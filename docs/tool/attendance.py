# -*- coding: utf-8 -*-
# @Author  : catnlp
# @FileName: statistics.py
# @Time    : 2020/9/1 23:26

import re
from openpyxl import load_workbook

import argparse

parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--excel', type=str, default="8月考勤.xlsx")
args = parser.parse_args()


class Attendance:
    """
    考勤类
    """
    def __init__(self, excel_file, save_file):
        self._excel = excel_file
        self._save = save_file
        self._attention_list = ["出勤天数", "上班缺卡次数", "下班缺卡次数",
                                "事假", "病假", "年假", "丧假", "陪产假", "产假", "婚假",
                                "旷工天数", "调休",
                                "入职时间", "离职时间",
                                "备注1", "备注2",
                                "上班迟到60到120分钟", "上班迟到30到60分钟",
                                "下班早退10到30分钟", "上班迟到120分钟以上",
                                "下班早退120分钟以上", "下班早退30到60分钟",
                                "下班早退0到10分钟", "下班早退60到120分钟",
                                "上班迟到0到10分钟", "上班迟到10到30分钟"]
        self.special = ["行政部", "人事部", "财务部", "法务部"]

    def process(self):
        wb = load_workbook(filename=self._excel)
        rs = wb['月度汇总']
        ws = wb.create_sheet('月度汇总_自动')

        search = {}
        pos2name = {}
        for i, line in enumerate(rs):
            print("i: ", i)
            # if (i + 1) % 100 == 0:
            #     print("i: ", i + 1)
            if i < 2 or i == 3:
                continue
            if i == 2:
                for j, col in enumerate(line):
                    if col.value not in search:
                        search[col.value] = j + 1
                    pos2name[j + 1] = col.value
                print(search)
                continue
            name = rs.cell(row=i + 1, column=search["姓名"]).value
            if not name:
                break

            remark1 = []
            remark2 = []
            entry = rs.cell(row=i + 1, column=search["入职时间"]).value
            resign = rs.cell(row=i + 1, column=search["离职时间"]).value
            if (entry == "#N/A" and resign == "#N/A") or \
                    (not entry and not resign):
                department = rs.cell(row=i + 1, column=search["部门"]).value
                if department in self.special:
                    # 统计请假情况
                    leave_dict = {}
                    holidays = ["事假", "病假", "年假", "丧假",
                                "陪产假", "产假", "婚假"]
                    for holiday in holidays:
                        if holiday not in search:
                            continue
                        personal_leave = rs.cell(row=i + 1,
                                                 column=search[holiday]).value
                        if personal_leave:
                            leave_dict[holiday] = personal_leave

                    # 调休
                    rest = rs.cell(row=i + 1, column=search["调休"]).value
                    if rest:
                        rest = float(rest)
                    left = search["调休"]
                    right = search["入职时间"]
                    ex_num = 0
                    for col in range(left, right):
                        cur = rs.cell(row=i + 1, column=col).value
                        cur = str(cur)
                        if isinstance(rest, float) and cur.find("休息并打卡") != -1:
                            if cur.find("4小时") != -1:
                                rest -= 0.5
                            else:
                                rest -= 1

                        elif cur.find("旷工") != -1 and pos2name[col] == "六":
                            ex_num += 1
                    if isinstance(rest, float) and rest > 0.4:
                        if "事假" not in leave_dict:
                            leave_dict["事假"] = rest
                        else:
                            leave_dict["事假"] += rest

                    if not leave_dict:
                        remark1.append("全勤")
                        remark2.append("全勤")
                    else:
                        tmp_list = []
                        tmp_list2 = []
                        for holiday in leave_dict:
                            day = leave_dict[holiday]
                            tmp_list.append(f"{holiday}{day}天")
                            tmp_list2.append(f"{day}天{holiday}")
                        remark1.append("，".join(tmp_list))
                        remark2.append("，".join(tmp_list2))

                    # 迟到统计
                    num = 2
                    lates = ["上班迟到0到10分钟", "上班迟到10到30分钟",
                             "上班迟到30到60分钟", "上班迟到60到120分钟",
                             "上班迟到120分钟以上"]
                    money_num = 0
                    money_day = 0
                    for late in lates:
                        if re.search(r"(\d+)到(\d+)", late):
                            match = re.search(r"(\d+)到(\d+)", late)
                            n1 = match.group(1)
                            n2 = match.group(2)
                            if n1 == '0':
                                num_late = rs.cell(row=i + 1,
                                                   column=search[late]).value
                                if num_late:
                                    num_late = int(num_late)
                                    if num_late > num:
                                        num_late -= num
                                        num = 0
                                        remark1.append(
                                            f"迟到{n1}-{n2}分钟{num_late}次")
                                        money_num += 10 * num_late
                                    else:
                                        num = max(num - num_late, 0)
                            else:
                                num_late = rs.cell(row=i + 1,
                                                   column=search[late]).value
                                if num_late:
                                    remark1.append(f"迟到{n1}-{n2}分钟{num_late}次")
                                    if n1 == "10":
                                        money_num += 20 * num_late
                                    elif n1 == "30":
                                        money_num += 30 * num_late
                                    else:
                                        money_day += 0.5 * num_late
                        elif re.search(r"\d+", late):
                            num_late = rs.cell(row=i + 1,
                                               column=search[late]).value
                            if num_late:
                                remark1.append(f"迟到120分钟以上{num_late}次")
                                money_day += num_late

                    earlys = ["下班早退0到10分钟", "下班早退10到30分钟",
                              "下班早退30到60分钟", "下班早退60到120分钟",
                              "下班早退120分钟以上", ]
                    for early in earlys:
                        if re.search(r"(\d+)到(\d+)", early):
                            match = re.search(r"(\d+)到(\d+)", early)
                            n1 = match.group(1)
                            n2 = match.group(2)
                            if n1 == '0':
                                num_early = rs.cell(row=i + 1,
                                                    column=search[early]).value
                                if num_early:
                                    num_early = int(num_early)
                                    if num_early > num:
                                        num_early -= num
                                        num = 0
                                        remark1.append(
                                            f"早退{n1}-{n2}分钟{num_early}次")
                                        money_num += 10 * num_early
                                    else:
                                        num = max(num - num_early, 0)
                            else:
                                num_early = rs.cell(row=i + 1,
                                                    column=search[early]).value
                                if num_early:
                                    remark1.append(f"早退{n1}-{n2}分钟{num_early}次")
                                    if n1 == "10":
                                        money_num += 20 * num_early
                                    elif n1 == "30":
                                        money_num += 30 * num_early
                                    else:
                                        money_day += 0.5 * num_early
                        elif re.search(r"\d+", early):
                            num_early = rs.cell(row=i + 1,
                                                column=search[early]).value
                            if num_early:
                                remark1.append(f"早退120分钟以上{num_early}次")
                                money_day += num_early
                    if money_num > 0 and money_day > 0.4:
                        remark2.append(f"迟到/早退扣{money_num}元+{money_day}天薪资")
                    elif money_num > 0:
                        remark2.append(f"迟到/早退扣{money_num}元")
                    elif money_day > 0.4:
                        remark2.append(f"迟到/早退扣{money_day}天薪资")
                    # 统计缺卡次数
                    missing_num = 0
                    missing_cards = ["上班缺卡次数", "下班缺卡次数"]
                    for missing_card in missing_cards:
                        num = rs.cell(row=i + 1,
                                      column=search[missing_card]).value
                        missing_num += int(num)
                    if missing_num > 0:
                        remark1.append(f"缺卡{missing_num}次")
                        remark2.append(f"缺卡{missing_num}次")

                    # 统计旷工
                    if "旷工天数" in search:
                        absenteeism = rs.cell(row=i + 1,
                                              column=search["旷工天数"]).value
                        absenteeism = int(absenteeism)
                        absenteeism = max(absenteeism - ex_num, 0)
                        if absenteeism > 0:
                            remark1.append(f"旷工{absenteeism}天")
                            remark2.append(f"旷工{absenteeism}天")
                else:
                    # 统计请假情况
                    leave_dict = {}
                    holidays = ["事假", "病假", "年假", "丧假",
                                "陪产假", "产假", "婚假"]
                    for holiday in holidays:
                        if holiday not in search:
                            continue
                        personal_leave = rs.cell(row=i + 1,
                                                 column=search[holiday]).value
                        if personal_leave:
                            leave_dict[holiday] = personal_leave

                    # 调休
                    rest = rs.cell(row=i + 1, column=search["调休"]).value
                    if rest:
                        rest = float(rest)
                    left = search["调休"]
                    right = search["入职时间"]
                    for col in range(left, right):
                        cur = rs.cell(row=i + 1, column=col).value
                        cur = str(cur)
                        if isinstance(rest, float) and cur.find("休息并打卡") != -1:
                            if cur.find("4小时") != -1:
                                rest -= 0.5
                            else:
                                rest -= 1
                    if isinstance(rest, float) and rest > 0.4:
                        if "事假" not in leave_dict:
                            leave_dict["事假"] = rest
                        else:
                            leave_dict["事假"] += rest

                    if not leave_dict:
                        remark1.append("全勤")
                        remark2.append("全勤")
                    else:
                        tmp_list = []
                        tmp_list2 = []
                        for holiday in leave_dict:
                            day = leave_dict[holiday]
                            tmp_list.append(f"{holiday}{day}天")
                            tmp_list2.append(f"{day}天{holiday}")
                        remark1.append("，".join(tmp_list))
                        remark2.append("，".join(tmp_list2))

                    # 迟到统计
                    num = 2
                    lates = ["上班迟到0到10分钟", "上班迟到10到30分钟",
                             "上班迟到30到60分钟", "上班迟到60到120分钟",
                             "上班迟到120分钟以上"]
                    money_num = 0
                    money_day = 0
                    for late in lates:
                        if re.search(r"(\d+)到(\d+)", late):
                            match = re.search(r"(\d+)到(\d+)", late)
                            n1 = match.group(1)
                            n2 = match.group(2)
                            if n1 == '0':
                                num_late = rs.cell(row=i + 1,
                                                   column=search[late]).value
                                if num_late:
                                    num_late = int(num_late)
                                    if num_late > num:
                                        num_late -= num
                                        num = 0
                                        remark1.append(
                                            f"迟到{n1}-{n2}分钟{num_late}次")
                                        money_num += 10 * num_late
                                    else:
                                        num = max(num - num_late, 0)
                            else:
                                num_late = rs.cell(row=i + 1,
                                                   column=search[late]).value
                                if num_late:
                                    remark1.append(f"迟到{n1}-{n2}分钟{num_late}次")
                                    if n1 == "10":
                                        money_num += 20 * num_late
                                    elif n1 == "30":
                                        money_num += 30 * num_late
                                    else:
                                        money_day += 0.5 * num_late
                        elif re.search(r"\d+", late):
                            num_late = rs.cell(row=i + 1,
                                               column=search[late]).value
                            if num_late:
                                remark1.append(f"迟到120分钟以上{num_late}次")
                                money_day += num_late

                    earlys = ["下班早退0到10分钟", "下班早退10到30分钟",
                              "下班早退30到60分钟", "下班早退60到120分钟",
                              "下班早退120分钟以上", ]
                    for early in earlys:
                        if re.search(r"(\d+)到(\d+)", early):
                            match = re.search(r"(\d+)到(\d+)", early)
                            n1 = match.group(1)
                            n2 = match.group(2)
                            if n1 == '0':
                                num_early = rs.cell(row=i + 1,
                                                    column=search[early]).value
                                if num_early:
                                    num_early = int(num_early)
                                    if num_early > num:
                                        num_early -= num
                                        num = 0
                                        remark1.append(
                                            f"早退{n1}-{n2}分钟{num_early}次")
                                        money_num += 10 * num_early
                                    else:
                                        num = max(num - num_early, 0)
                            else:
                                num_early = rs.cell(row=i + 1,
                                                    column=search[early]).value
                                if num_early:
                                    remark1.append(f"早退{n1}-{n2}分钟{num_early}次")
                                    if n1 == "10":
                                        money_num += 20 * num_early
                                    elif n1 == "30":
                                        money_num += 30 * num_early
                                    else:
                                        money_day += 0.5 * num_early
                        elif re.search(r"\d+", early):
                            num_early = rs.cell(row=i + 1,
                                                column=search[early]).value
                            if num_early:
                                remark1.append(f"早退120分钟以上{num_early}次")
                                money_day += num_early
                    if money_num > 0 and money_day > 0.4:
                        remark2.append(f"迟到/早退扣{money_num}元+{money_day}天薪资")
                    elif money_num > 0:
                        remark2.append(f"迟到/早退扣{money_num}元")
                    elif money_day > 0.4:
                        remark2.append(f"迟到/早退扣{money_day}天薪资")
                    # 统计缺卡次数
                    missing_num = 0
                    missing_cards = ["上班缺卡次数", "下班缺卡次数"]
                    for missing_card in missing_cards:
                        num = rs.cell(row=i + 1, column=search[missing_card]).value
                        missing_num += int(num)
                    if missing_num > 0:
                        remark1.append(f"缺卡{missing_num}次")
                        remark2.append(f"缺卡{missing_num}次")

                    # 统计旷工
                    if "旷工天数" in search:
                        absenteeism = rs.cell(row=i + 1, column=search["旷工天数"]).value
                        if absenteeism > 0:
                            remark1.append(f"旷工{absenteeism}天")
                            remark2.append(f"旷工{absenteeism}天")

            else:
                # 出勤天数
                days = rs.cell(row=i + 1, column=search["出勤天数"]).value
                remark1.append(f"共上班{days}天")
                remark2.append(f"共上班{days}天")
                # 统计请假情况
                leave_dict = {}
                holidays = ["病假", "年假", "丧假",
                            "陪产假", "产假", "婚假"]
                for holiday in holidays:
                    if holiday not in search:
                        continue
                    personal_leave = rs.cell(row=i + 1,
                                             column=search[holiday]).value
                    if personal_leave:
                        leave_dict[holiday] = personal_leave

                if leave_dict:
                    tmp_list = []
                    tmp_list2 = []
                    for holiday in leave_dict:
                        day = leave_dict[holiday]
                        tmp_list.append(f"{holiday}{day}天")
                        tmp_list2.append(f"{day}天{holiday}")
                    remark1.append("，".join(tmp_list))
                    remark2.append("，".join(tmp_list2))

                # 迟到统计
                num = 2
                lates = ["上班迟到0到10分钟", "上班迟到10到30分钟",
                         "上班迟到30到60分钟", "上班迟到60到120分钟",
                         "上班迟到120分钟以上"]
                money_num = 0
                money_day = 0
                for late in lates:
                    if re.search(r"(\d+)到(\d+)", late):
                        match = re.search(r"(\d+)到(\d+)", late)
                        n1 = match.group(1)
                        n2 = match.group(2)
                        if n1 == '0':
                            num_late = rs.cell(row=i + 1,
                                               column=search[late]).value
                            if num_late:
                                num_late = int(num_late)
                                if num_late > num:
                                    num_late -= num
                                    num = 0
                                    remark1.append(
                                        f"迟到{n1}-{n2}分钟{num_late}次")
                                    money_num += 10 * num_late
                                else:
                                    num = max(num - num_late, 0)
                        else:
                            num_late = rs.cell(row=i + 1,
                                               column=search[late]).value
                            if num_late:
                                remark1.append(f"迟到{n1}-{n2}分钟{num_late}次")
                                if n1 == "10":
                                    money_num += 20 * num_late
                                elif n1 == "30":
                                    money_num += 30 * num_late
                                else:
                                    money_day += 0.5 * num_late
                    elif re.search(r"\d+", late):
                        num_late = rs.cell(row=i + 1,
                                           column=search[late]).value
                        if num_late:
                            remark1.append(f"迟到120分钟以上{num_late}次")
                            money_day += num_late

                earlys = ["下班早退0到10分钟", "下班早退10到30分钟",
                          "下班早退30到60分钟", "下班早退60到120分钟",
                          "下班早退120分钟以上", ]
                for early in earlys:
                    if re.search(r"(\d+)到(\d+)", early):
                        match = re.search(r"(\d+)到(\d+)", early)
                        n1 = match.group(1)
                        n2 = match.group(2)
                        if n1 == '0':
                            num_early = rs.cell(row=i + 1,
                                                column=search[early]).value
                            if num_early:
                                num_early = int(num_early)
                                if num_early > num:
                                    num_early -= num
                                    num = 0
                                    remark1.append(
                                        f"早退{n1}-{n2}分钟{num_early}次")
                                    money_num += 10 * num_early
                                else:
                                    num = max(num - num_early, 0)
                        else:
                            num_early = rs.cell(row=i + 1,
                                                column=search[early]).value
                            if num_early:
                                remark1.append(f"早退{n1}-{n2}分钟{num_early}次")
                                if n1 == "10":
                                    money_num += 20 * num_early
                                elif n1 == "30":
                                    money_num += 30 * num_early
                                else:
                                    money_day += 0.5 * num_early
                    elif re.search(r"\d+", early):
                        num_early = rs.cell(row=i + 1,
                                            column=search[early]).value
                        if num_early:
                            remark1.append(f"早退120分钟以上{num_early}次")
                            money_day += num_early
                if money_num > 0 and money_day > 0.4:
                    remark2.append(f"迟到/早退扣{money_num}元+{money_day}天薪资")
                elif money_num > 0:
                    remark2.append(f"迟到/早退扣{money_num}元")
                elif money_day > 0.4:
                    remark2.append(f"迟到/早退扣{money_day}天薪资")

                # 统计缺卡次数
                missing_num = 0
                missing_cards = ["上班缺卡次数", "下班缺卡次数"]
                for missing_card in missing_cards:
                    num = rs.cell(row=i + 1, column=search[missing_card]).value
                    missing_num += int(num)
                if missing_num > 0:
                    remark1.append(f"缺卡{missing_num}次")
                    remark2.append(f"缺卡{missing_num}次")

            ws.cell(row=i, column=1).value = rs.cell(row=i + 1,
                                                     column=search["备注1"]).value
            ws.cell(row=i, column=3).value = rs.cell(row=i + 1,
                                                     column=search["备注2"]).value
            ws.cell(row=i, column=2).value = "，".join(remark1).replace(".0", "")
            ws.cell(row=i, column=4).value = "，".join(remark2).replace(".0", "")

            # if i == 4:
            #     entry = rs.cell(row=i + 1, column=search["入职时间"]).value
            #     print(entry)
            #     if entry == "#N/A":
            #         print("yes")
            #     resign = rs.cell(row=i + 1, column=search["离职时间"]).value
            #     print(resign)
            #     print(str(resign).split(" ")[0].split("-")[-1])
            #     break
        wb.save(self._save)


if __name__ == "__main__":
    attend = Attendance(args.excel, args.excel.replace("考勤", "考勤_自动"))
    attend.process()
